import os, re, json, pathlib, datetime
import pdfplumber, pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ---------- CONFIG ----------
PDF_DIR = pathlib.Path("hindalco_pdfs")            # where your workflow saves Hindalco PDFs
DATA_DIR = pathlib.Path("data")
DATA_DIR.mkdir(parents=True, exist_ok=True)

EXCEL_FILE = DATA_DIR / "hindalco_prices.xlsx"
LAST_PROCESSED_FILE = DATA_DIR / "last_hindalco_processed.txt"   # stores last processed PDF filename
LATEST_JSON = pathlib.Path("latest_hindalco_pdf.json")           # created by your existing downloader (if present)

COLUMNS = ["Sl.no.", "Description", "Grade", "Basic Price", "Circular Date", "Circular Link"]

# match date like "...-08-august-2025.pdf" or "...-8-august-2025.pdf"
FILENAME_DATE_RE = re.compile(r"(\d{1,2})[-_ ]([A-Za-z]+)[-_ ](\d{4})", re.IGNORECASE)

# ---------- HELPERS ----------
def newest_pdf(pdf_dir: pathlib.Path) -> pathlib.Path | None:
    if not pdf_dir.exists():
        return None
    pdfs = sorted(pdf_dir.glob("*.pdf"), key=lambda p: p.stat().st_mtime, reverse=True)
    return pdfs[0] if pdfs else None

def latest_link_from_json() -> str | None:
    if LATEST_JSON.exists():
        try:
            obj = json.loads(LATEST_JSON.read_text(encoding="utf-8"))
            for k in ("pdf_url", "url", "latest_pdf_url"):
                if k in obj and isinstance(obj[k], str) and obj[k].lower().endswith(".pdf"):
                    return obj[k]
        except Exception:
            pass
    return None

def parse_date_from_filename(name: str) -> str:
    """
    Extract date like '08.08.2025' from a filename having '...-08-august-2025...' (Hindalco style).
    Fallback: today's date.
    """
    m = FILENAME_DATE_RE.search(name)
    if not m:
        return datetime.date.today().strftime("%d.%m.%Y")
    day, mon_text, year = m.groups()
    day = int(day); year = int(year)
    months = {m: i for i, m in enumerate(
        ["january","february","march","april","may","june","july","august","september","october","november","december"], start=1)}
    mm = months.get(mon_text.lower())
    if not mm:
        return datetime.date.today().strftime("%d.%m.%Y")
    try:
        d = datetime.date(year, mm, day)
        return d.strftime("%d.%m.%Y")
    except ValueError:
        return datetime.date.today().strftime("%d.%m.%Y")

def divide_thousands(value: str | float | int) -> float | None:
    s = str(value).replace(",", "").strip()
    if not s:
        return None
    try:
        return round(float(s) / 1000.0, 3)
    except ValueError:
        return None

def extract_row_from_pdf(pdf_path: pathlib.Path) -> tuple[str, str]:
    """
    Returns (description, raw_price).
    Look for a line/row that contains all of: P0610, P1020, 'EC Grade'.
    Robust to table or text layout.
    """
    must_have = ["P0610", "P1020", "EC GRADE"]
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # 1) Try tables first
            try:
                tables = page.extract_tables()
            except Exception:
                tables = []
            for tbl in tables or []:
                for row in tbl:
                    cells = [(c or "").strip() for c in row]
                    line = " ".join(cells).upper()
                    if all(x in line for x in must_have):
                        # price: last numeric-looking cell
                        price = ""
                        for c in reversed(cells):
                            if re.search(r"\d", c):
                                price = c.replace(",", "").strip()
                                break
                        desc = " ".join(cells[:-1]).strip()
                        return desc, price
            # 2) Fallback: text flow lines
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
            lines = {}
            for w in words:
                y = round(w["top"], 1)
                lines.setdefault(y, []).append(w)
            for _, wlist in lines.items():
                text_line = " ".join([w["text"] for w in sorted(wlist, key=lambda x: x["x0"])])
                u = text_line.upper()
                if all(x in u for x in must_have):
                    m = re.search(r"(\d{5,7}(?:\.\d+)?)\s*$", text_line.replace(",", ""))
                    price = m.group(1) if m else ""
                    desc = text_line.strip()
                    return desc, price
    raise RuntimeError("Could not find the Hindalco target row (P0610 / P1020 / EC Grade).")

def sort_df(df: pd.DataFrame) -> pd.DataFrame:
    dtd = pd.to_datetime(df["Circular Date"], dayfirst=True, errors="coerce")
    df = df.assign(_d=dtd).sort_values(by=["_d", "Sl.no."], ascending=[False, True], kind="stable").drop(columns=["_d"])
    df["Basic Price"] = pd.to_numeric(df["Basic Price"], errors="coerce").round(3)
    df["Circular Date"] = pd.to_datetime(df["Circular Date"], dayfirst=True, errors="coerce").dt.strftime("%d.%m.%Y")
    return df

def save_excel_formatted(df: pd.DataFrame, path: pathlib.Path):
    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active

    center = Alignment(horizontal="center", vertical="center")
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for v in df[col_name].astype(str).values:
            max_len = max(max_len, len(v))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 80))

    header_row = 1
    link_col_idx = COLUMNS.index("Circular Link") + 1
    price_col_idx = COLUMNS.index("Basic Price") + 1

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            if r > header_row and c == price_col_idx:
                cell.number_format = "0.000"
            if r > header_row and c == link_col_idx:
                val = cell.value
                if isinstance(val, str) and val.startswith("http"):
                    cell.hyperlink = val

    ws.freeze_panes = "A2"
    wb.save(path)

# ---------- MAIN ----------
def main():
    pdf_path = newest_pdf(PDF_DIR)
    if not pdf_path:
        print("No PDFs found in hindalco_pdfs/ — run the downloader first.", flush=True)
        return

    # avoid duplicates
    last_name = LAST_PROCESSED_FILE.read_text(encoding="utf-8").strip() if LAST_PROCESSED_FILE.exists() else ""
    if pdf_path.name == last_name:
        print("Latest PDF already processed. Skipping Excel update.", flush=True)
        return

    desc, raw_price = extract_row_from_pdf(pdf_path)
    price = divide_thousands(raw_price)
    if price is None:
        raise RuntimeError(f"Could not parse numeric price: {raw_price!r}")

    # date (dd.mm.yyyy) from filename
    circ_date = parse_date_from_filename(pdf_path.name)

    # Grade: per your instruction, always P1020
    grade = "P1020"

    # link from JSON (if your downloader writes it)
    link = latest_link_from_json() or ""

    # read or create excel
    if EXCEL_FILE.exists():
        df = pd.read_excel(EXCEL_FILE, dtype={"Sl.no.":"Int64"})
        for c in COLUMNS:
            if c not in df.columns:
                df[c] = pd.NA
        df = df[COLUMNS]
        next_slno = int(df["Sl.no."].max()) + 1 if df["Sl.no."].notna().any() else 1
    else:
        df = pd.DataFrame(columns=COLUMNS)
        next_slno = 1

    new_row = {
        "Sl.no.": next_slno,
        "Description": desc,
        "Grade": grade,
        "Basic Price": price,
        "Circular Date": circ_date,
        "Circular Link": link,
    }

    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df = sort_df(df)
    save_excel_formatted(df, EXCEL_FILE)

    LAST_PROCESSED_FILE.write_text(pdf_path.name, encoding="utf-8")
    print(f"Appended 1 row. Excel updated at {EXCEL_FILE}")

if __name__ == "__main__":
    main()
