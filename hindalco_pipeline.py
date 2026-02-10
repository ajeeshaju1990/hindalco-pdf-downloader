import os
import re
import json
import pathlib
import datetime
import argparse
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


# -------------------- CONFIG --------------------
START_URL = "https://www.hindalco.com/businesses/aluminium/primary-aluminium/primary-metal-price"

PDF_DIR = pathlib.Path("hindalco_pdfs")
DATA_DIR = pathlib.Path("data")
PDF_DIR.mkdir(parents=True, exist_ok=True)
DATA_DIR.mkdir(parents=True, exist_ok=True)

LATEST_JSON = pathlib.Path("latest_hindalco_pdf.json")
LAST_PROCESSED_FILE = DATA_DIR / "last_hindalco_processed.txt"
PROCESSED_SET_FILE = DATA_DIR / "processed_files.txt"
EXCEL_FILE = DATA_DIR / "hindalco_prices.xlsx"

# ✅ NEW: manual overrides file (you will maintain this when needed)
OVERRIDES_FILE = DATA_DIR / "hindalco_manual_overrides.xlsx"

DAILY_COLUMNS = ["Date", "Description", "Grade", "Basic Price", "Circular Date", "Circular Link"]

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

# --- filename date: "07-february-2026" etc ---
FILENAME_DATE_RE = re.compile(
    r"(?i)(\d{1,2})[\-_ ]+(january|february|march|april|may|june|july|august|"
    r"september|october|november|december)[\-_ ]+(\d{4})"
)

# --- flexible date in Excel: dd.mm.yyyy OR dd-mm-yyyy OR dd/mm/yyyy or embedded in text ---
DATE_ANY_RE = re.compile(r"(\d{1,2})[./-](\d{1,2})[./-](\d{4})")
URL_RE = re.compile(r"(https?://\S+)")


# -------------------- HELPERS --------------------
def _normalize_url(u: str) -> str:
    try:
        p = urlparse(u)
        return p._replace(query="", fragment="").geturl()
    except Exception:
        return u


def _extract_date_any(value) -> datetime.date | None:
    """Parse dd.mm.yyyy / dd-mm-yyyy / dd/mm/yyyy from messy cells."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    s = str(value).strip()
    m = DATE_ANY_RE.search(s)
    if not m:
        return None
    dd, mm, yyyy = m.group(1), m.group(2), m.group(3)
    try:
        return datetime.date(int(yyyy), int(mm), int(dd))
    except ValueError:
        return None


def _extract_first_url(*vals) -> str:
    for v in vals:
        if v is None:
            continue
        if isinstance(v, float) and pd.isna(v):
            continue
        m = URL_RE.search(str(v))
        if m:
            return m.group(1)
    return ""


def _month_name(d: datetime.date) -> str:
    return [
        "january","february","march","april","may","june",
        "july","august","september","october","november","december"
    ][d.month - 1]


def guess_hindalco_pdf_url(cdate: datetime.date) -> str:
    """Generate expected Hindalco PDF URL for a circular date."""
    dd = f"{cdate.day:02d}"
    month = _month_name(cdate)
    yyyy = str(cdate.year)
    return f"https://www.hindalco.com/Upload/PDF/primary-ready-reckoner-{dd}-{month}-{yyyy}.pdf"


def ensure_event_link(event: dict) -> dict:
    """If event link missing, generate one from circular date."""
    if not event.get("clink"):
        event["clink"] = guess_hindalco_pdf_url(event["cdate"])
    return event


# -------------------- HTML & DOWNLOAD --------------------
def get_html(url: str) -> str:
    r = requests.get(url, headers={"User-Agent": UA}, timeout=60)
    r.raise_for_status()
    return r.text


def find_latest_pdf_url(html: str) -> str | None:
    soup = BeautifulSoup(html, "html.parser")
    anchors = soup.find_all("a", href=True)
    candidates: list[tuple[int, str]] = []

    for a in anchors:
        href = (a["href"] or "").strip()
        if not href.lower().endswith(".pdf"):
            continue
        abs_url = urljoin(START_URL, href)
        text = (a.get_text(" ", strip=True) or "").lower()

        score = 0
        if any(k in text for k in ("ready", "reckoner", "price")):
            score += 2
        if any(k in href.lower() for k in ("ready", "reckoner", "price")):
            score += 2
        candidates.append((score, abs_url))

    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def read_latest_json() -> dict:
    if LATEST_JSON.exists():
        try:
            return json.loads(LATEST_JSON.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def write_latest_json(pdf_url: str, filename: str):
    obj = {
        "last_pdf_url": pdf_url,
        "download_timestamp": datetime.datetime.utcnow().isoformat() + "Z",
        "filename": filename,
    }
    LATEST_JSON.write_text(json.dumps(obj, indent=2), encoding="utf-8")


def download_pdf(pdf_url: str) -> pathlib.Path:
    headers = {
        "User-Agent": UA,
        "Accept": "application/pdf,*/*;q=0.9",
        "Referer": START_URL,
    }
    with requests.get(pdf_url, headers=headers, timeout=60, stream=True, allow_redirects=True) as r:
        r.raise_for_status()
        ctype = (r.headers.get("Content-Type", "") or "").lower()
        if "application/pdf" not in ctype:
            raise RuntimeError(f"Expected PDF but got Content-Type={ctype!r}")

        name = os.path.basename(urlparse(r.url).path)
        timestamp = datetime.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        fname = f"{timestamp}_{name}" if name else f"hindalco_{timestamp}.pdf"

        dest = PDF_DIR / fname
        with open(dest, "wb") as f:
            for chunk in r.iter_content(65536):
                if chunk:
                    f.write(chunk)
        return dest


# -------------------- DATE FROM FILENAME --------------------
def parse_date_from_filename(filename_or_url: str) -> datetime.date | None:
    s = filename_or_url or ""
    m = FILENAME_DATE_RE.search(s)
    if not m:
        return None
    day, mon_text, year = m.groups()
    months = {
        "january": 1, "february": 2, "march": 3, "april": 4,
        "may": 5, "june": 6, "july": 7, "august": 8,
        "september": 9, "october": 10, "november": 11, "december": 12,
    }
    mm = months.get(mon_text.lower())
    if not mm:
        return None
    try:
        return datetime.date(int(year), mm, int(day))
    except ValueError:
        return None


# -------------------- PDF PARSE --------------------
def divide_thousands(x: str | float | int) -> float | None:
    s = str(x).replace(",", "").strip()
    if not s:
        return None
    try:
        return round(float(s) / 1000.0, 3)
    except ValueError:
        return None


def extract_target_row(pdf_path: pathlib.Path) -> tuple[str, str]:
    must_have = ["P0610", "P1020", "EC GRADE"]

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            try:
                tables = page.extract_tables()
            except Exception:
                tables = []
            for tbl in tables or []:
                for row in tbl:
                    cells = [(c or "").strip() for c in row]
                    line = " ".join(cells).upper()
                    if all(x in line for x in must_have):
                        price = ""
                        for c in reversed(cells):
                            if re.search(r"\d", c):
                                price = c.replace(",", "").strip()
                                break
                        desc = " ".join(cells[:-1]).strip()
                        return desc, price

        words = pdf.pages[0].extract_words(use_text_flow=True, keep_blank_chars=False) if pdf.pages else []
        lines: dict[float, list[dict]] = {}
        for w in words:
            y = round(w["top"], 1)
            lines.setdefault(y, []).append(w)

        for _, wlist in lines.items():
            text_line = " ".join([w["text"] for w in sorted(wlist, key=lambda x: x["x0"])])
            u = text_line.upper()
            if all(x in u for x in must_have):
                m = re.search(r"(\d{5,7}(?:\.\d+)?)\s*$", text_line.replace(",", ""))
                price = m.group(1) if m else ""
                desc = text_line.strip()
                return desc, price

    raise RuntimeError(f"Could not find target row in: {pdf_path.name}")


# -------------------- OVERRIDES --------------------
def load_manual_overrides() -> list[dict]:
    """
    Reads data/hindalco_manual_overrides.xlsx if present.

    Expected columns (any subset works):
      - Circular Date (mandatory for a row)
      - Circular Link (optional)
      - Basic Price (optional)
      - Description (optional)
      - Grade (optional)
    """
    if not OVERRIDES_FILE.exists():
        return []

    df = pd.read_excel(OVERRIDES_FILE)

    # normalize column names
    df.columns = [str(c).strip() for c in df.columns]

    # allow common variants
    col_map = {
        "Circular Date": "Circular Date",
        "Circular Link": "Circular Link",
        "Basic Price": "Basic Price",
        "Description": "Description",
        "Grade": "Grade",
    }

    # Must have circular date
    if "Circular Date" not in df.columns:
        return []

    overrides = []
    for _, r in df.iterrows():
        cdt = _extract_date_any(r.get("Circular Date"))
        if not cdt:
            continue
        overrides.append({
            "cdate": cdt,
            "clink": (r.get("Circular Link") if isinstance(r.get("Circular Link"), str) else "") or "",
            "price": pd.to_numeric(r.get("Basic Price"), errors="coerce"),
            "desc": (r.get("Description") if r.get("Description") is not None else "") or "",
            "grade": (r.get("Grade") if r.get("Grade") is not None else "") or "",
        })
    return overrides


def apply_overrides(events: list[dict], overrides: list[dict]) -> list[dict]:
    """Merge overrides into events by circular date, without deleting existing data."""
    by_date = {e["cdate"]: e for e in events}

    for o in overrides:
        cdate = o["cdate"]
        if cdate not in by_date:
            # create minimal event if missing
            by_date[cdate] = {
                "desc": o.get("desc", ""),
                "grade": o.get("grade") or "P1020",
                "price": float(o["price"]) if pd.notna(o.get("price")) else None,
                "cdate": cdate,
                "clink": o.get("clink", ""),
            }
        else:
            e = by_date[cdate]
            # only override if provided
            if o.get("desc"):
                e["desc"] = o["desc"]
            if o.get("grade"):
                e["grade"] = o["grade"]
            if pd.notna(o.get("price")):
                e["price"] = float(o["price"])
            if o.get("clink"):
                e["clink"] = o["clink"]

    # ensure links exist (auto-generate if still empty)
    out = []
    for e in by_date.values():
        out.append(ensure_event_link(e))
    out.sort(key=lambda x: x["cdate"])
    return out


# -------------------- EVENTS from EXCEL --------------------
def load_events_from_excel_if_any() -> list[dict]:
    if not EXCEL_FILE.exists():
        return []

    df = pd.read_excel(EXCEL_FILE)
    df.columns = [str(c).strip() for c in df.columns]

    for k in ["Description", "Grade", "Basic Price", "Circular Date", "Circular Link"]:
        if k not in df.columns:
            df[k] = pd.NA

    # Parse circular date flexibly
    df["Circular Date DT"] = df["Circular Date"].apply(_extract_date_any)
    # if still missing, try parse from Circular Link field (some sheets have the date there)
    miss = df["Circular Date DT"].isna()
    if miss.any():
        df.loc[miss, "Circular Date DT"] = df.loc[miss, "Circular Link"].apply(_extract_date_any)

    df = df.dropna(subset=["Circular Date DT"])
    df["Basic Price"] = pd.to_numeric(df["Basic Price"], errors="coerce")

    # Clean link column: if not http, try extract from any cell
    df["Circular Link"] = df.apply(
        lambda r: r["Circular Link"] if isinstance(r["Circular Link"], str) and r["Circular Link"].startswith("http")
        else _extract_first_url(r["Circular Link"], r["Circular Date"], r["Description"]),
        axis=1
    )

    # Dedupe events by circular date (keep last)
    ev = df.sort_values(by="Circular Date DT").drop_duplicates(subset=["Circular Date DT"], keep="last")

    events = []
    for _, r in ev.iterrows():
        cdate = r["Circular Date DT"]
        link = r.get("Circular Link", "") or ""
        events.append({
            "desc": r.get("Description", "") or "",
            "grade": r.get("Grade", "P1020") or "P1020",
            "price": float(r.get("Basic Price")) if pd.notna(r.get("Basic Price")) else None,
            "cdate": cdate,
            "clink": link,
        })

    # Auto-fill links if missing
    events = [ensure_event_link(e) for e in events]
    events.sort(key=lambda e: e["cdate"])
    return events


# -------------------- BUILD DAILY --------------------
def _fmt_date_dash(d: datetime.date) -> str:
    return d.strftime("%d-%m-%Y")


def _fmt_date_dot(d: datetime.date) -> str:
    return d.strftime("%d.%m.%Y")


def build_daily_from_events(events: list[dict], end_date: datetime.date | None = None) -> pd.DataFrame:
    if not events:
        return pd.DataFrame(columns=DAILY_COLUMNS)

    events = sorted(events, key=lambda e: e["cdate"])
    start = events[0]["cdate"]
    today = end_date or datetime.date.today()
    if start > today:
        start = today

    rows = []
    idx = 0
    current = events[0]

    for d in (start + datetime.timedelta(n) for n in range((today - start).days + 1)):
        while idx + 1 < len(events) and events[idx + 1]["cdate"] <= d:
            idx += 1
            current = events[idx]

        rows.append({
            "Date": _fmt_date_dash(d),
            "Description": current["desc"],
            "Grade": current["grade"] or "P1020",
            "Basic Price": round(float(current["price"]), 3) if current["price"] is not None else None,
            "Circular Date": _fmt_date_dot(current["cdate"]),
            "Circular Link": current["clink"],
        })

    df = pd.DataFrame(rows)
    df["DateDT"] = pd.to_datetime(df["Date"], format="%d-%m-%Y", errors="coerce")
    df = df.sort_values(by="DateDT", ascending=False).drop(columns=["DateDT"])
    return df[DAILY_COLUMNS]


# -------------------- WRITE EXCEL --------------------
def save_excel_formatted(df: pd.DataFrame, path: pathlib.Path):
    if df is None or df.empty:
        pd.DataFrame(columns=DAILY_COLUMNS).to_excel(path, index=False)
        return

    df.to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    center = Alignment(horizontal="center", vertical="center")

    for cidx, cname in enumerate(df.columns, start=1):
        max_len = len(str(cname))
        for v in df[cname].tolist():
            try:
                sv = "" if (v is None or pd.isna(v)) else str(v)
            except Exception:
                sv = str(v)
            max_len = max(max_len, len(sv))
        ws.column_dimensions[get_column_letter(cidx)].width = max(12, min(max_len + 2, 80))

    header_row = 1
    price_col_idx = DAILY_COLUMNS.index("Basic Price") + 1
    link_col_idx = DAILY_COLUMNS.index("Circular Link") + 1

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            if r > header_row and c == price_col_idx:
                cell.number_format = "0.000"

        if r > header_row:
            val = ws.cell(row=r, column=link_col_idx).value
            if isinstance(val, str) and val.startswith("http"):
                ws.cell(row=r, column=link_col_idx).hyperlink = val

    ws.freeze_panes = "A2"
    wb.save(path)


# -------------------- STATE FILES --------------------
def load_processed_set() -> set[str]:
    if PROCESSED_SET_FILE.exists():
        return set(x.strip() for x in PROCESSED_SET_FILE.read_text(encoding="utf-8").splitlines() if x.strip())
    return set()


def save_processed_set(s: set[str]):
    PROCESSED_SET_FILE.write_text("\n".join(sorted(s)), encoding="utf-8")


# -------------------- MODES --------------------
def run_normal():
    events = load_events_from_excel_if_any()

    # apply overrides each scheduled run (your manual safety net)
    overrides = load_manual_overrides()
    if overrides:
        events = apply_overrides(events, overrides)

    html = get_html(START_URL)
    pdf_url = find_latest_pdf_url(html)

    latest = read_latest_json()
    last_url = latest.get("last_pdf_url", "")

    if pdf_url and _normalize_url(pdf_url) != _normalize_url(last_url):
        pdf_path = download_pdf(pdf_url)
        write_latest_json(pdf_url, str(pdf_path))
        print(f"Downloaded: {pdf_path.name}")

        # parse and add event
        desc, raw_price = extract_target_row(pdf_path)
        price = divide_thousands(raw_price)
        if price is None:
            raise RuntimeError(f"Could not parse numeric price: {raw_price!r}")

        cdate = parse_date_from_filename(pdf_path.name) or parse_date_from_filename(pdf_url) or datetime.date.today()

        # merge event (keep existing link if new is empty)
        # always prefer real URL from download
        new_event = {"desc": desc, "grade": "P1020", "price": price, "cdate": cdate, "clink": pdf_url}
        events = [e for e in events if e["cdate"] != cdate]
        events.append(new_event)
        events.sort(key=lambda e: e["cdate"])

        LAST_PROCESSED_FILE.write_text(pdf_path.name, encoding="utf-8")
        processed = load_processed_set()
        processed.add(pdf_path.name)
        save_processed_set(processed)

        print(f"Added/updated event for {cdate.strftime('%d.%m.%Y')}")
    else:
        print("No new circular; will forward-fill daily series.")

    daily_df = build_daily_from_events(events, end_date=datetime.date.today())
    save_excel_formatted(daily_df, EXCEL_FILE)
    print(f"Wrote daily sheet with {len(daily_df)} rows → {EXCEL_FILE}")


def run_backfill():
    """
    Backfill parses PDFs in hindalco_pdfs/ and creates/updates events.
    Links for manual PDFs are auto-generated if unknown.
    Overrides are applied at the end.
    """
    pdfs = sorted(PDF_DIR.glob("*.pdf"), key=lambda p: p.stat().st_mtime)
    events = load_events_from_excel_if_any()
    processed = load_processed_set()

    # quick index
    by_date = {e["cdate"]: e for e in events}

    added = 0
    for pdf_path in pdfs:
        cdate = parse_date_from_filename(pdf_path.name)
        if not cdate:
            continue

        # If already processed AND already have an event for that date, skip
        if pdf_path.name in processed and cdate in by_date:
            continue

        try:
            desc, raw_price = extract_target_row(pdf_path)
            price = divide_thousands(raw_price)
            if price is None:
                print(f"Could not parse price in {pdf_path.name}; skipping.")
                continue

            # if manual pdf: generate expected URL
            link = by_date.get(cdate, {}).get("clink", "") if cdate in by_date else ""
            if not link:
                link = guess_hindalco_pdf_url(cdate)

            by_date[cdate] = {
                "desc": desc,
                "grade": "P1020",
                "price": price,
                "cdate": cdate,
                "clink": link,
            }

            processed.add(pdf_path.name)
            added += 1

        except Exception as e:
            print(f"Error processing {pdf_path.name}: {e}")

    save_processed_set(processed)

    events = list(by_date.values())
    events = [ensure_event_link(e) for e in events]
    events.sort(key=lambda e: e["cdate"])

    # apply overrides at end
    overrides = load_manual_overrides()
    if overrides:
        events = apply_overrides(events, overrides)

    daily_df = build_daily_from_events(events, end_date=datetime.date.today())
    save_excel_formatted(daily_df, EXCEL_FILE)

    if added == 0:
        print("Backfill complete. No new events added.")
    else:
        print(f"Backfill complete. Added {added} event(s). Rebuilt daily sheet with {len(daily_df)} rows.")


def run_repair():
    events = load_events_from_excel_if_any()

    # apply overrides always
    overrides = load_manual_overrides()
    if overrides:
        events = apply_overrides(events, overrides)

    if not events:
        print("No events present to rebuild from.")
        return

    daily_df = build_daily_from_events(events, end_date=datetime.date.today())
    save_excel_formatted(daily_df, EXCEL_FILE)
    print(f"Repair complete. Rebuilt daily sheet with {len(daily_df)} rows.")


# -------------------- ENTRYPOINT --------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--backfill", default="false", help="true/false (process all existing PDFs once)")
    ap.add_argument("--repair", default="false", help="true/false (rebuild daily sheet from existing data)")
    args = ap.parse_args()

    if str(args.repair).strip().lower() in ("true", "1", "yes", "y"):
        run_repair()
    elif str(args.backfill).strip().lower() in ("true", "1", "yes", "y"):
        run_backfill()
    else:
        run_normal()


if __name__ == "__main__":
    main()
